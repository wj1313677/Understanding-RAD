import os
from glob import glob
import openpyxl
import pandas as pd

def remove_strikethrough_text(ws):
    cleaned_data = []
    for row in ws.iter_rows():
        cleaned_row = []
        for cell in row:
            if cell.value is not None:
                # Remove value if strikethrough is True
                if cell.font is not None and getattr(cell.font, 'strike', False):
                    cleaned_row.append('')
                else:
                    cleaned_row.append(cell.value)
            else:
                cleaned_row.append('')
        cleaned_data.append(cleaned_row)
    return cleaned_data

xlsx_files = glob('**/*.xlsx', recursive=True)

for xlsx in xlsx_files:
    wb = openpyxl.load_workbook(xlsx, data_only=True)
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        cleaned_data = remove_strikethrough_text(ws)
        if not cleaned_data:
            continue
        # Convert to DataFrame
        df = pd.DataFrame(cleaned_data)
        # Find the "Change Ind." column (case-insensitive)
        header_row = df.iloc[0].astype(str).str.lower()
        change_ind_col = None
        for idx, col_name in enumerate(header_row):
            if col_name.strip() == "change ind.":
                change_ind_col = idx
                break
        # If found, filter rows and drop column
        if change_ind_col is not None:
            # Remove rows with "SUP" or "DEL" in the Change Ind. column (excluding the header)
            mask = ~df.iloc[1:, change_ind_col].astype(str).str.strip().str.upper().isin(["SUP", "DEL"])
            # Keep header and filtered rows
            df = pd.concat([df.iloc[[0]], df.iloc[1:][mask]])
            # Drop the "Change Ind." column
            df = df.drop(df.columns[change_ind_col], axis=1)
            df.reset_index(drop=True, inplace=True)
        # Build CSV file name
        basename = os.path.splitext(xlsx)[0]
        csv_name = f"{basename}_{sheet_name}.csv"
        csv_name = csv_name.replace('/', '_').replace('\\', '_').replace(' ', '_')
        df.to_csv(csv_name, index=False, header=False)
        print(f"Converted {xlsx} [{sheet_name}] -> {csv_name}")
