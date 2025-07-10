import os
import glob
import openpyxl
import pandas as pd
from openpyxl.cell.rich_text import CellRichText

def process_excel_file(xlsx_file_path):
    print(f"Processing {xlsx_file_path}...")
    
    # Load the workbook with rich_text=True to access formatting
    try:
        wb_full = openpyxl.load_workbook(xlsx_file_path, rich_text=True)
        wb_data_only = openpyxl.load_workbook(xlsx_file_path, data_only=True)
    except FileNotFoundError:
        print(f"Error: The file {xlsx_file_path} was not found.")
        return

    for sheet_name in wb_full.sheetnames:
        
        ws_data_only = wb_data_only[sheet_name]
        ws_full = wb_full[sheet_name]

        # Get all rows with formatting and data_only values
        all_rows_full = list(ws_full.iter_rows())
        all_rows_data_only = list(ws_data_only.iter_rows(values_only=True))

        if not all_rows_full: # Check if sheet is empty
            print(f"Sheet {sheet_name} is empty. Skipping.")
            continue

        # Create a temporary DataFrame from data_only values to easily find "Change Ind."
        df_temp_for_filtering = pd.DataFrame(all_rows_data_only)

        # Use the first column (index 0) for filtering and removal
        change_ind_col_idx = 0

        # Step 1: Identify rows to keep based on the first column
        rows_to_keep_indices = []
        rows_to_keep_indices.append(0) # Always keep header

        for r_idx in range(1, len(all_rows_data_only)): # Start from 1 to skip header
            cell_value = str(df_temp_for_filtering.iloc[r_idx, change_ind_col_idx]).strip().upper()
            if cell_value not in ["SUP", "DEL"]:
                rows_to_keep_indices.append(r_idx)
            else:
                pass

        # Step 2: Process only the kept rows for strikethrough removal
        final_processed_data = []
        for r_idx in rows_to_keep_indices:
            original_row_cells = all_rows_full[r_idx]
            current_row_cleaned_values = []

            for c_idx, cell_orig in enumerate(original_row_cells):
                # Apply strikethrough removal logic
                processed_value = cell_orig.value
                if isinstance(cell_orig.value, CellRichText):
                    new_text_parts = []
                    for text_part in cell_orig.value:
                        # Ensure we get the text from the rich text part
                        part_text = text_part.text if hasattr(text_part, 'text') else str(text_part)
                        if not (hasattr(text_part, 'font') and text_part.font and text_part.font.strike):
                            new_text_parts.append(part_text)
                    if new_text_parts:
                        processed_value = "".join(new_text_parts)
                    else:
                        processed_value = '' # Remove entirely if all parts are strikethrough
                elif cell_orig.font and cell_orig.font.strike:
                    processed_value = '' # Remove entirely if whole cell is strikethrough
                
                current_row_cleaned_values.append(processed_value)
            final_processed_data.append(current_row_cleaned_values)

        if not final_processed_data:
            print(f"Sheet {sheet_name} resulted in no data after processing. Skipping CSV creation.")
            continue

        # Convert to DataFrame for final column dropping
        df_final = pd.DataFrame(final_processed_data)

        # Step 3: Drop the first 3 column (which was used for filtering)
        df_final = df_final.drop(df_final.columns[[0, 1, 2]], axis=1)
        df_final.reset_index(drop=True, inplace=True)

        # Construct CSV file name using only the worksheet name
        csv_name = f"{sheet_name}.csv"
        csv_name = csv_name.replace(' ', '_') # Replace spaces with underscores

        df_final.to_csv(csv_name, index=False, header=False)
        print(f"Converted {xlsx_file_path} [{sheet_name}] -> {csv_name}")

# Process all .xlsx files in the current directory
xlsx_files = glob.glob("*.xlsx")
if not xlsx_files:
    print("No .xlsx files found in the current directory.")
else:
    for excel_file in xlsx_files:
        process_excel_file(excel_file)